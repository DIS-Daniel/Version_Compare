# app.py
import streamlit as st
import paramiko
import stat
import difflib
import posixpath
from typing import List, Tuple
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import io
import fnmatch

# ------------------------
# Page config
# ------------------------
st.set_page_config(
    page_title="GK Version Comparison",
    page_icon="🧾",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.title("🧾 GK Version Comparison")
st.caption(
    "Compare text files line-by-line between Old and New SFTP folders and download a color-coded Excel report. "
    "Preview only changed lines per file and see live comparison progress."
)

# ------------------------
# Helper functions
# ------------------------
def file_name_only(path: str) -> str:
    return posixpath.basename(path.rstrip("/"))

def file_diff_status(file_status: str, diff_lines):
    if file_status == "Side-by-side diff":
        return "Changed" if any(ls != "Unchanged" for _, _, ls in diff_lines) else "No differences"
    if file_status == "Only in Old Folder":
        return "Only in Old"
    if file_status == "Only in New Folder":
        return "Only in New"
    return file_status

# ------------------------
# SFTP helpers
# ------------------------
def connect_sftp(host, port, username, password):
    transport = paramiko.Transport((host, port))
    transport.connect(username=username, password=password)
    sftp = paramiko.SFTPClient.from_transport(transport)
    return sftp, transport

def list_files(sftp, path):
    files = []
    path = path.rstrip("/")
    try:
        for entry in sftp.listdir_attr(path):
            full = posixpath.join(path, entry.filename)
            if stat.S_ISDIR(entry.st_mode):
                files.extend(list_files(sftp, full))
            else:
                files.append(full)
    except Exception:
        pass
    return files

def read_file_lines(sftp, filepath):
    try:
        with sftp.open(filepath, "rb") as f:
            return f.read().decode(errors="ignore").splitlines()
    except Exception:
        return None

# ------------------------
# Diff logic
# ------------------------
def side_by_side_diff(lines_old, lines_new):
    diff = []
    sm = difflib.SequenceMatcher(None, lines_old, lines_new)
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        old_chunk = lines_old[i1:i2]
        new_chunk = lines_new[j1:j2]
        max_len = max(len(old_chunk), len(new_chunk))
        old_chunk += [""] * (max_len - len(old_chunk))
        new_chunk += [""] * (max_len - len(new_chunk))
        for o, n in zip(old_chunk, new_chunk):
            status = (
                "Unchanged" if tag == "equal" else
                "Modified" if tag == "replace" else
                "Removed" if tag == "delete" else
                "Added"
            )
            diff.append((o, n, status))
    return diff

def rel_map(files, root):
    root = root.rstrip("/")
    return {f[len(root)+1:]: f for f in files if f.startswith(root + "/")}

def should_ignore(rel_path, patterns):
    return any(fnmatch.fnmatch(rel_path, p) for p in patterns)

def compare_folders(sftp_old, sftp_new, folder_old, folder_new, ignore_patterns, hide_ignored, progress_callback=None):
    files_old = rel_map(list_files(sftp_old, folder_old), folder_old)
    files_new = rel_map(list_files(sftp_new, folder_new), folder_new)

    all_files = sorted(set(files_old) | set(files_new))
    diffs = []

    total = len(all_files)
    for i, rel in enumerate(all_files, 1):
        po = files_old.get(rel)
        pn = files_new.get(rel)

        if should_ignore(rel, ignore_patterns):
            if not hide_ignored:
                diffs.append((rel, "Ignored (pattern)", []))
            if progress_callback:
                progress_callback(i, total, rel)
            continue

        if po and not pn:
            diffs.append((rel, "Only in Old Folder", []))
        elif pn and not po:
            diffs.append((rel, "Only in New Folder", []))
        else:
            lo = read_file_lines(sftp_old, po)
            ln = read_file_lines(sftp_new, pn)
            if lo is None or ln is None:
                diffs.append((rel, "Binary or unreadable file", []))
            else:
                diffs.append((rel, "Side-by-side diff", side_by_side_diff(lo, ln)))

        if progress_callback:
            progress_callback(i, total, rel)

    return diffs

# ------------------------
# Excel export
# ------------------------
def generate_excel_bytes(differences, old_label, new_label, only_diffs=True):
    wb = Workbook()
    ws = wb.active

    ws.append([f"GK Version Comparison — Old: {old_label or 'N/A'} | New: {new_label or 'N/A'}"])
    ws.merge_cells("A1:D1")
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([])

    ws.append(["File", old_label or "Old Line", new_label or "New Line", "Status"])
    for c in range(1,5):
        ws.cell(row=3, column=c).font = Font(bold=True)

    fills = {
        "Added": PatternFill("solid", fgColor="C6EFCE"),
        "Removed": PatternFill("solid", fgColor="FFC7CE"),
        "Modified": PatternFill("solid", fgColor="FFEB9C"),
    }

    for f, status, lines in differences:
        if status in {"Binary or unreadable file", "Ignored (pattern)"}:
            continue

        fname = file_name_only(f)

        if status == "Side-by-side diff":
            for o, n, s in lines:
                if only_diffs and s == "Unchanged":
                    continue
                ws.append([fname, o, n, s])
                if s in fills:
                    ws.cell(row=ws.max_row, column=2).fill = fills[s]
                    ws.cell(row=ws.max_row, column=3).fill = fills[s]
        else:
            ws.append([fname, "", "", file_diff_status(status, [])])

    for i in range(1,5):
        ws.column_dimensions[get_column_letter(i)].width = 50

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

# ------------------------
# Sidebar inputs
# ------------------------
with st.sidebar:
    st.header("Connection Settings")
    same_server = st.checkbox("Use the same server for Old and New folders", value=True)

    with st.form("controls_form", clear_on_submit=False):
        st.subheader("Old Version")
        host_old = st.text_input("Host", key="host_old")
        port_old = st.number_input("Port", 1, 65535, 22, key="port_old")
        user_old = st.text_input("Username", key="user_old")
        pass_old = st.text_input("Password", type="password", key="pass_old")
        folder_old = st.text_input("Folder Path", key="folder_old")

        st.subheader("New Version")
        host_new = st.text_input("Host", key="host_new", disabled=same_server)
        port_new = st.number_input("Port", 1, 65535, 22, key="port_new", disabled=same_server)
        user_new = st.text_input("Username", key="user_new", disabled=same_server)
        pass_new = st.text_input("Password", type="password", key="pass_new", disabled=same_server)
        folder_new = st.text_input("Folder Path", key="folder_new")

        st.divider()
        old_label = st.text_input("Old Version Label", key="old_label")
        new_label = st.text_input("New Version Label", key="new_label")
        only_diffs_excel = st.checkbox("Excel: only include differences", True)

        submitted = st.form_submit_button("Run Comparison")

# ------------------------
# Run comparison
# ------------------------
if submitted:
    if same_server:
        host_new, port_new, user_new, pass_new = host_old, port_old, user_old, pass_old

    progress_bar = st.progress(0)
    status_text = st.empty()

    def update_progress(processed, total, current_file):
        pct = int(processed / max(total,1) * 100)
        progress_bar.progress(pct)
        status_text.text(f"Processing {processed}/{total}: {current_file}")

    sftp_old = sftp_new = t_old = t_new = None
    try:
        with st.spinner("Connecting to server(s)..."):
            sftp_old, t_old = connect_sftp(host_old, port_old, user_old, pass_old)
            if same_server:
                sftp_new, t_new = sftp_old, t_old
            else:
                sftp_new, t_new = connect_sftp(host_new, port_new, user_new, pass_new)

        with st.spinner("Comparing folders..."):
            default_ignores = ["*.png","*.jpg","*.jpeg","*.gif","*.bmp","*.pdf","*.zip","*.gz","*.tar","*.7z","*.docx","*.xlsx","*.pptx"]
            differences = compare_folders(
                sftp_old, sftp_new,
                folder_old, folder_new,
                ignore_patterns=default_ignores,
                hide_ignored=True,
                progress_callback=update_progress
            )

        st.session_state["differences"] = differences
        st.session_state["only_diffs_excel"] = only_diffs_excel
        st.success("Comparison complete!")

    finally:
        if not same_server and sftp_new: sftp_new.close()
        if sftp_old: sftp_old.close()
        if t_old: t_old.close()
        if t_new and not same_server: t_new.close()

# ------------------------
# Display results, metrics, preview, Excel
# ------------------------
if "differences" in st.session_state:
    differences = st.session_state["differences"]
    only_diffs_excel = st.session_state.get("only_diffs_excel", True)

    # --- Compute rows & metrics ---
    computed_rows = [
        {"File": file_name_only(f), "Status": file_diff_status(s,l), "RawStatus": s,
         "FullPath": f, "DiffLines": l, "HasChanges": any(ls!="Unchanged" for _,_,ls in l)}
        for f,s,l in differences if s not in {"Binary or unreadable file","Ignored (pattern)"}
    ]
    total_changed = sum(1 for r in computed_rows if r["Status"]=="Changed")
    total_only_old = sum(1 for r in computed_rows if r["Status"]=="Only in Old")
    total_only_new = sum(1 for r in computed_rows if r["Status"]=="Only in New")
    total_text_diffed = sum(1 for r in computed_rows if r["RawStatus"]=="Side-by-side diff")

    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Changed", total_changed)
    c2.metric("Only in Old", total_only_old)
    c3.metric("Only in New", total_only_new)
    c4.metric("Text files diffed", total_text_diffed)

    # --- File-level filters ---
    st.subheader("File-level results")
    all_statuses = ["Changed","No differences","Only in Old","Only in New"]
    selected_statuses = st.multiselect("Filter by status", options=all_statuses, default=["Changed","Only in Old","Only in New"])
    file_search = st.text_input("Search file names (contains)")
    has_changes_filter = st.checkbox("Only show files with changes", value=False)

    filtered_rows = computed_rows
    if selected_statuses:
        filtered_rows = [r for r in filtered_rows if r["Status"] in selected_statuses]
    if file_search:
        filtered_rows = [r for r in filtered_rows if file_search.lower() in r["File"].lower()]
    if has_changes_filter:
        filtered_rows = [r for r in filtered_rows if r.get("HasChanges")]

    st.dataframe([{"File": r["File"], "Status": r["Status"]} for r in filtered_rows], use_container_width=True)

    # --- Preview ---
    preview_candidates = [r for r in filtered_rows if r["RawStatus"]=="Side-by-side diff"]
    if preview_candidates:
        st.subheader("Preview changed lines")
        sel_file = st.selectbox("Choose file", [r["File"] for r in preview_candidates], key="preview_file")
        sel_row = next(r for r in preview_candidates if r["File"]==sel_file)
        changed_lines = [(o,n,s) for o,n,s in sel_row["DiffLines"] if s!="Unchanged"]

        if changed_lines:
            MAX_PREVIEW = 500
            preview_rows = changed_lines[:MAX_PREVIEW]
            st.caption(f"Showing first {len(preview_rows)} of {len(changed_lines)} changed lines")
            st.dataframe([{"Old":o,"New":n,"Status":s} for o,n,s in preview_rows], use_container_width=True)
        else:
            st.info("No changed lines to preview.")

    # --- Excel download with optional filter ---
    st.subheader("Download Excel report")
    apply_filter_to_excel = st.checkbox("Apply file status filter to Excel export", value=False)
    if apply_filter_to_excel:
        keep_paths = {r["FullPath"] for r in filtered_rows}
        export_differences = [(f,s,l) for f,s,l in differences if f in keep_paths]
    else:
        export_differences = differences

    excel_bytes = generate_excel_bytes(export_differences, old_label, new_label, only_diffs_excel)
    st.download_button(
        "📥 Download Excel",
        excel_bytes,
        "side_by_side_report.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
